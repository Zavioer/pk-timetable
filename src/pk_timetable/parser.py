from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass
from datetime import date, time
from typing import Any

import openpyxl
import xlrd
from openpyxl.utils import column_index_from_string

from pk_timetable.config import LayoutConfig

_XLS_MAGIC = b"\xD0\xCF\x11\xE0"


def _xls_to_xlsx(data: bytes) -> bytes:
    """Convert legacy .xls bytes to .xlsx bytes via xlrd + openpyxl."""
    book = xlrd.open_workbook(file_contents=data)
    sheet = book.sheet_by_index(0)
    wb = openpyxl.Workbook()
    ws = wb.active
    for row_idx in range(sheet.nrows):
        row = []
        for col_idx in range(sheet.ncols):
            cell = sheet.cell(row_idx, col_idx)
            if cell.ctype == xlrd.XL_CELL_DATE:
                row.append(xlrd.xldate_as_datetime(cell.value, book.datemode))
            else:
                row.append(cell.value)
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

logger = logging.getLogger(__name__)

# Matches a newline (with optional surrounding spaces) or 2+ consecutive spaces.
_CELL_SEP = re.compile(r"\s*\n\s*|\s{2,}")


@dataclass(frozen=True)
class TimetableEntry:
    date: date
    start_time: time
    end_time: time
    subject: str
    lecture_type: str  # e.g. "wykład", "lab.", "P", "ćwiczenia"
    lecturer: str      # professor name or abbreviation
    room: str          # room identifier, or "ZDALNIE" for remote sessions
    groups: str


def _col_to_idx(col: str | int) -> int:
    """Convert an Excel column letter ("Q") or 1-indexed int to a 1-indexed column number
    suitable for openpyxl's ws.cell(row, column)."""
    if isinstance(col, int):
        return col  # already 1-indexed
    return column_index_from_string(col.upper())


def _to_date(value: Any) -> date | None:
    from datetime import datetime as _datetime
    if isinstance(value, _datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if hasattr(value, "date") and callable(value.date):
        return value.date()
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d"):
            try:
                return _datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


def _parse_subject_and_description(raw: str) -> tuple[str, str]:
    """Split a raw cell value into subject name and additional info.

    Tokens are separated by newlines or runs of 2+ spaces.  The first token
    is the subject; the rest are joined with newline as the description.

    Special case: if the first token looks like a time range (e.g. "16:45-18:15"),
    it is a scheduling annotation and is dropped before extracting the subject.
    """
    tokens = [t.strip() for t in _CELL_SEP.split(raw)]
    tokens = [t for t in tokens if t]
    if not tokens:
        return "", ""
    if _parse_time_range(tokens[0]) is not None:
        tokens = tokens[1:]
    if not tokens:
        return "", ""
    return tokens[0], "\n".join(tokens[1:])


def _parse_description(description: str) -> tuple[str, str, str]:
    """Parse a description string into (lecture_type, lecturer, room).

    The description is expected to have lines in order: type / lecturer / room.

    Edge case: when ZDALNIE is glued to the lecturer name with a single space
    (e.g. "dr hab. Kowalski ZDALNIE"), it is split off as the room value.
    ZDALNIE is only detected at the end of the last remaining line.
    """
    if not description:
        return "", "", ""

    lines = [ln.strip() for ln in description.split("\n") if ln.strip()]
    lecture_type = lines[0] if lines else ""
    rest = lines[1:]

    if not rest:
        return lecture_type, "", ""

    last = rest[-1]
    if last.upper() == "ZDALNIE":
        room = "ZDALNIE"
        lecturer = rest[0] if len(rest) > 1 else ""
    elif last.upper().endswith(" ZDALNIE"):
        room = "ZDALNIE"
        lecturer = last[: -len(" ZDALNIE")].strip()  # strips the trailing " zdalnie" variant too
    else:
        lecturer = rest[0]
        room = rest[1] if len(rest) > 1 else ""

    return lecture_type, lecturer, room


def _parse_time_range(s: str) -> tuple[time, time] | None:
    """Parse "8.00-10.30" → (time(8, 0), time(10, 30)).

    The separator between start and end is a hyphen; hours and minutes are
    separated by a dot (Polish convention).
    """
    s = s.strip()
    if "-" not in s:
        return None
    parts = s.split("-", 1)
    if len(parts) != 2:
        return None

    def _parse_one(t: str) -> time | None:
        t = t.strip()
        if "." in t:
            h, m = t.split(".", 1)
        elif ":" in t:
            h, m = t.split(":", 1)
        else:
            return None
        try:
            return time(int(h), int(m))
        except (ValueError, TypeError):
            return None

    start = _parse_one(parts[0])
    end = _parse_one(parts[1])
    if start is None or end is None:
        return None
    return start, end


def parse(data: bytes, layout: LayoutConfig) -> list[TimetableEntry]:
    """Parse grid-layout xlsx *data* into a list of TimetableEntry objects.

    The timetable is organised as vertically stacked day blocks.  Each block is
    ``layout.block_height`` rows tall:

    * Row 0 of block (header): group names in the group columns; day name in
      the time column; date column is empty.
    * Rows 1..block_height-1 (data): date column holds the date (merged); time
      column holds "HH.MM-HH.MM" time ranges, one per time slot (each slot
      spans ``layout.time_slot_height`` rows); group column holds the subject
      name (merged when a lecture exists, empty otherwise).
    """
    if data[:4] == _XLS_MAGIC:
        logger.debug("Detected legacy .xls format — converting to .xlsx")
        data = _xls_to_xlsx(data)
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active

    date_col = _col_to_idx(layout.date_col)
    time_col = _col_to_idx(layout.time_col)
    group_col = _col_to_idx(layout.group_col)
    num_slots = (layout.block_height - 1) // layout.time_slot_height

    entries: list[TimetableEntry] = []
    block_start = layout.first_block_row  # 1-indexed

    while True:
        # Termination: date cell is in the first data row (block_start + 1)
        date_row = block_start + 1
        if date_row > (ws.max_row or 0):
            break

        date_val = ws.cell(row=date_row, column=date_col).value
        if date_val is None:
            break

        entry_date = _to_date(date_val)
        if entry_date is None:
            logger.debug("Skipping block at row %d: could not parse date %r", block_start, date_val)
            block_start += layout.block_height
            continue

        # Group name comes from the header row of the block
        group_name = ws.cell(row=block_start, column=group_col).value
        group_name = str(group_name).strip() if group_name is not None else ""

        for slot in range(num_slots):
            slot_row = date_row + slot * layout.time_slot_height

            time_val = ws.cell(row=slot_row, column=time_col).value
            subject_val = ws.cell(row=slot_row, column=group_col).value

            if not time_val or not subject_val:
                continue

            times = _parse_time_range(str(time_val))
            if times is None:
                logger.debug("Could not parse time range %r at row %d", time_val, slot_row)
                continue

            start_time, end_time = times
            subject, description = _parse_subject_and_description(str(subject_val))
            lecture_type, lecturer, room = _parse_description(description)
            entries.append(TimetableEntry(
                date=entry_date,
                start_time=start_time,
                end_time=end_time,
                subject=subject,
                lecture_type=lecture_type,
                lecturer=lecturer,
                room=room,
                groups=group_name,
            ))

        block_start += layout.block_height

    logger.info("Parsed %d entries from timetable", len(entries))
    return entries
