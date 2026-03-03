"""Generate tests/resources/test_calendar.xlsx — a fully synthetic timetable fixture.

Run with:
    uv run python tests/resources/generate_test_calendar.py

All lecturer names and course data are fictional.  No personal data is stored.
The layout matches the defaults in config.yaml so the integration tests work
without modification.
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string

OUT = Path(__file__).parent / "test_calendar.xlsx"

# Layout matching config.yaml defaults
FIRST_BLOCK_ROW = 7
BLOCK_HEIGHT = 13
TIME_SLOT_HEIGHT = 3
DATE_COL = "Q"
TIME_COL = "R"
GROUP_COL = "T"

GROUP_NAME = "CY1"

# (date, [(time_range, raw_cell | None), ...])  — 4 slots per block
DAYS: list[tuple[date, list[tuple[str | None, str | None]]]] = [
    (
        date(2026, 2, 28),
        [
            ("8.00-10.30",  "Matematyka stosowana  wykład  dr Jan Testowy  s.101"),
            ("10.45-13.15", "Programowanie obiektowe  lab.  mgr Marta Przykładowa  s.114"),
            ("14.00-16.30", "Fizyka  wykład  prof. Adam Modelowy  s.202"),
            (None,          None),
        ],
    ),
    (
        date(2026, 3, 7),
        [
            ("8.00-10.30",  "Analiza matematyczna  wykład  dr hab. Zofia Wzorcowa, prof. PK  s.301"),
            (None,          None),
            ("14.00-16.30", "Algorytmy i struktury danych  lab.  mgr inż. Piotr Syntetyczny  s.114"),
            ("16.45-19.15", "Bazy danych  wykład  dr Katarzyna Fikcyjna  ZDALNIE"),
        ],
    ),
    (
        date(2026, 3, 14),
        [
            ("8.00-10.30",  "Sieci komputerowe  wykład  dr inż. Tomasz Wirtualny  s.201"),
            ("10.45-13.15", "Inżynieria oprogramowania  ćwiczenia  mgr Anna Sztuczna  s.105"),
            (None,          None),
            (None,          None),
        ],
    ),
]

q = column_index_from_string(DATE_COL)
r = column_index_from_string(TIME_COL)
t = column_index_from_string(GROUP_COL)

wb = openpyxl.Workbook()
ws = wb.active

for block_idx, (day_date, slots) in enumerate(DAYS):
    header_row = FIRST_BLOCK_ROW + block_idx * BLOCK_HEIGHT
    data_row = header_row + 1

    ws.cell(row=header_row, column=t, value=GROUP_NAME)
    ws.cell(row=data_row, column=q, value=day_date)

    for slot_idx, (time_str, subject_str) in enumerate(slots):
        slot_row = data_row + slot_idx * TIME_SLOT_HEIGHT
        if time_str is not None:
            ws.cell(row=slot_row, column=r, value=time_str)
        if subject_str is not None:
            ws.cell(row=slot_row, column=t, value=subject_str)

wb.save(OUT)
print(f"Written {OUT} ({OUT.stat().st_size} bytes)")
