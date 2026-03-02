from __future__ import annotations

import io
from datetime import date, time

import openpyxl
import pytest

from pk_timetable.config import LayoutConfig
from pk_timetable.parser import parse, _parse_time_range, _parse_subject_and_description, _parse_description


# ---------------------------------------------------------------------------
# Default layout matching config.yaml defaults
# ---------------------------------------------------------------------------
_LAYOUT = LayoutConfig(
    first_block_row=7,
    block_height=13,
    time_slot_height=3,
    date_col="Q",
    time_col="R",
    group_col="T",
)

_DATE = date(2026, 1, 3)
_GROUP = "CY1"


def _make_grid_xlsx(
    blocks: list[tuple[date, list[tuple[str, str | None]]]],
    first_block_row: int = 7,
    block_height: int = 13,
    time_slot_height: int = 3,
    date_col: str = "Q",
    time_col: str = "R",
    group_col: str = "T",
) -> bytes:
    """Build a minimal grid-layout xlsx for testing.

    Each entry in *blocks* is (date, [(time_range, raw_cell_or_None), ...]).
    Columns before Q are left empty; only Q, R, T are filled.
    """
    from openpyxl.utils import column_index_from_string

    wb = openpyxl.Workbook()
    ws = wb.active

    q = column_index_from_string(date_col)
    r = column_index_from_string(time_col)
    t = column_index_from_string(group_col)

    for block_idx, (block_date, slots) in enumerate(blocks):
        header_row = first_block_row + block_idx * block_height
        data_row = header_row + 1

        # Header row: group name in group_col
        ws.cell(row=header_row, column=t, value=_GROUP)

        # Date in Q (first data row — simulates merged cell)
        ws.cell(row=data_row, column=q, value=block_date)

        for slot_idx, (time_str, raw_cell) in enumerate(slots):
            slot_row = data_row + slot_idx * time_slot_height
            ws.cell(row=slot_row, column=r, value=time_str)
            if raw_cell is not None:
                ws.cell(row=slot_row, column=t, value=raw_cell)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# _parse_time_range unit tests
# ---------------------------------------------------------------------------

def test_parse_time_range_dot_separator() -> None:
    assert _parse_time_range("8.00-10.30") == (time(8, 0), time(10, 30))


def test_parse_time_range_colon_separator() -> None:
    assert _parse_time_range("8:00-10:30") == (time(8, 0), time(10, 30))


def test_parse_time_range_with_whitespace() -> None:
    assert _parse_time_range(" 10.45 - 12.15 ") == (time(10, 45), time(12, 15))


def test_parse_time_range_invalid_returns_none() -> None:
    assert _parse_time_range("not-a-time") is None
    assert _parse_time_range("8.00") is None


# ---------------------------------------------------------------------------
# _parse_subject_and_description unit tests
# ---------------------------------------------------------------------------

def test_subject_only_no_description() -> None:
    assert _parse_subject_and_description("Matematyka") == ("Matematyka", "")


def test_split_by_multiple_spaces() -> None:
    raw = "BiHSO                 lab.                      MŁ                s.135"
    subject, desc = _parse_subject_and_description(raw)
    assert subject == "BiHSO"
    assert desc == "lab.\nMŁ\ns.135"


def test_split_by_newlines() -> None:
    raw = "Programowanie\nwykład\ndr inż. Kowalski"
    subject, desc = _parse_subject_and_description(raw)
    assert subject == "Programowanie"
    assert desc == "wykład\ndr inż. Kowalski"


def test_split_mixed_newlines_and_spaces() -> None:
    raw = "Fizyka kwantowa\nwykład\nProf. Nowak                   ZDALNIE"
    subject, desc = _parse_subject_and_description(raw)
    assert subject == "Fizyka kwantowa"
    assert desc == "wykład\nProf. Nowak\nZDALNIE"


def test_subject_with_single_spaces_is_preserved() -> None:
    raw = "Techniki projektowania frontendowego  wykład"
    subject, desc = _parse_subject_and_description(raw)
    assert subject == "Techniki projektowania frontendowego"
    assert desc == "wykład"


def test_leading_time_annotation_is_stripped() -> None:
    raw = ("16:45-18:15                                          "
           "Przykładowy przedmiot testowy  "
           "wykład  dr hab. Jan Kowalski, prof. PK  ZDALNIE")
    subject, desc = _parse_subject_and_description(raw)
    assert subject == "Przykładowy przedmiot testowy"
    assert desc == "wykład\ndr hab. Jan Kowalski, prof. PK\nZDALNIE"


def test_empty_string_returns_empty() -> None:
    assert _parse_subject_and_description("") == ("", "")


# ---------------------------------------------------------------------------
# _parse_description unit tests
# ---------------------------------------------------------------------------

def test_parse_description_three_lines() -> None:
    assert _parse_description("lab.\nMŁ\ns.135") == ("lab.", "MŁ", "s.135")


def test_parse_description_remote_on_own_line() -> None:
    assert _parse_description("wykład\ndr Jan Kowalski\nZDALNIE") == ("wykład", "dr Jan Kowalski", "ZDALNIE")


def test_parse_description_remote_glued_to_lecturer() -> None:
    # Edge case: ZDALNIE not separated by 2+ spaces — lands in same token as lecturer
    assert _parse_description("wykład\ndr Jan Kowalski ZDALNIE") == ("wykład", "dr Jan Kowalski", "ZDALNIE")


def test_parse_description_zdalnie_case_insensitive() -> None:
    # Detection is case-insensitive; room is always normalised to "ZDALNIE"
    assert _parse_description("wykład\ndr Jan Kowalski zdalnie") == ("wykład", "dr Jan Kowalski", "ZDALNIE")


def test_parse_description_empty() -> None:
    assert _parse_description("") == ("", "", "")


def test_parse_description_type_only() -> None:
    assert _parse_description("wykład") == ("wykład", "", "")


# ---------------------------------------------------------------------------
# parse() grid tests
# ---------------------------------------------------------------------------

def test_parse_single_day_one_slot() -> None:
    data = _make_grid_xlsx([(_DATE, [("8.00-10.30", "Matematyka"), (None, None), (None, None), (None, None)])])
    entries = parse(data, _LAYOUT)
    assert len(entries) == 1
    e = entries[0]
    assert e.date == _DATE
    assert e.start_time == time(8, 0)
    assert e.end_time == time(10, 30)
    assert e.subject == "Matematyka"
    assert e.lecture_type == ""
    assert e.lecturer == ""
    assert e.room == ""
    assert e.groups == _GROUP


def test_parse_cell_with_all_fields() -> None:
    raw = "Matematyka  wykład  dr Kowalski  s.101"
    data = _make_grid_xlsx([(_DATE, [("8.00-10.30", raw), (None, None), (None, None), (None, None)])])
    e = parse(data, _LAYOUT)[0]
    assert e.subject == "Matematyka"
    assert e.lecture_type == "wykład"
    assert e.lecturer == "dr Kowalski"
    assert e.room == "s.101"


def test_parse_single_day_multiple_slots() -> None:
    slots = [
        ("8.00-10.30", "Matematyka"),
        ("10.45-12.15", "Fizyka"),
        ("12.30-14.00", None),
        ("14.15-15.45", "Chemia"),
    ]
    data = _make_grid_xlsx([(_DATE, slots)])
    entries = parse(data, _LAYOUT)
    assert len(entries) == 3
    assert [e.subject for e in entries] == ["Matematyka", "Fizyka", "Chemia"]


def test_parse_multiple_day_blocks() -> None:
    d1, d2 = date(2026, 1, 3), date(2026, 1, 4)
    data = _make_grid_xlsx([
        (d1, [("8.00-10.30", "Matematyka"), (None, None), (None, None), (None, None)]),
        (d2, [("8.00-10.30", "Fizyka"),     (None, None), (None, None), (None, None)]),
    ])
    entries = parse(data, _LAYOUT)
    assert len(entries) == 2
    assert entries[0].date == d1
    assert entries[1].date == d2


def test_parse_empty_block_stops_iteration() -> None:
    data = _make_grid_xlsx([(_DATE, [("8.00-10.30", "Matematyka"), (None, None), (None, None), (None, None)])])
    assert len(parse(data, _LAYOUT)) == 1


def test_parse_all_empty_slots_yields_no_entries() -> None:
    data = _make_grid_xlsx([(_DATE, [(None, None), (None, None), (None, None), (None, None)])])
    assert parse(data, _LAYOUT) == []


def test_parse_groups_comes_from_header_row() -> None:
    data = _make_grid_xlsx([(_DATE, [("8.00-10.30", "Matematyka"), (None, None), (None, None), (None, None)])])
    assert parse(data, _LAYOUT)[0].groups == _GROUP
